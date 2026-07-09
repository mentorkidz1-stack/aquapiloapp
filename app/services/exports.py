"""Exports Excel (openpyxl) et PDF (reportlab) des points de période."""
import io
from datetime import date as date_type, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

from app.services.rapports import (detail_produits, indicateurs_periode,
                                   repartition_paiements)

BLEU = "1F4E79"


def _fmt_fcfa(n) -> str:
    return f"{int(n):,}".replace(",", " ") + " FCFA"


def _jours_de(debut, fin):
    j = debut
    while j <= fin:
        yield j
        j += timedelta(days=1)


def titre_periode(type_p: str, debut: date_type, fin: date_type) -> str:
    if type_p == "jour":
        return f"Point journalier — {debut.strftime('%d/%m/%Y')}"
    if type_p == "semaine":
        return (f"Point hebdomadaire — du {debut.strftime('%d/%m/%Y')} "
                f"au {fin.strftime('%d/%m/%Y')}")
    return f"Point mensuel — {debut.strftime('%m/%Y')}"


# ================= EXCEL =================
def export_excel(type_p: str, debut: date_type, fin: date_type,
                 boutique_id: int | None = None,
                 boutique_nom: str = "Toutes boutiques") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"

    gras = Font(bold=True)
    blanc_gras = Font(bold=True, color="FFFFFF")
    fond_bleu = PatternFill("solid", fgColor=BLEU)
    bordure = Border(*[Side(style="thin", color="B0B0B0")] * 4)
    droite = Alignment(horizontal="right")

    # Titre
    ws["A1"] = "POISSONNERIE DONA — Cotonou"
    ws["A1"].font = Font(bold=True, size=14, color=BLEU)
    ws["A2"] = titre_periode(type_p, debut, fin) + f" — {boutique_nom}"
    ws["A2"].font = Font(bold=True, size=12)

    # Indicateurs
    ind = indicateurs_periode(debut, fin, boutique_id)
    lignes_ind = [("Chiffre d'affaires", _fmt_fcfa(ind["ca"])),
                  ("Marge", _fmt_fcfa(ind["marge"])),
                  ("Nombre de ventes", ind["nb_ventes"]),
                  ("Volume vendu", f"{ind['volume']:g}"),
                  ("Pertes (valeur)", _fmt_fcfa(ind["pertes"]))]
    r = 4
    for libelle, valeur in lignes_ind:
        ws.cell(row=r, column=1, value=libelle).font = gras
        c = ws.cell(row=r, column=2, value=valeur)
        c.alignment = droite
        r += 1

    # Synthèse par jour (semaine / mois)
    if type_p != "jour":
        r += 1
        entetes = ["Date", "CA", "Marge", "Ventes", "Pertes"]
        for i, e in enumerate(entetes, start=1):
            c = ws.cell(row=r, column=i, value=e)
            c.font, c.fill, c.border = blanc_gras, fond_bleu, bordure
        r += 1
        for j in _jours_de(debut, fin):
            ij = indicateurs_periode(j, j, boutique_id)
            valeurs = [j.strftime("%d/%m/%Y"), ij["ca"], ij["marge"],
                       ij["nb_ventes"], ij["pertes"]]
            for i, v in enumerate(valeurs, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = bordure
                if i > 1:
                    c.number_format = "#,##0"
                    c.alignment = droite
            r += 1

    # Feuille détail par produit
    ws2 = wb.create_sheet("Détail produits")
    entetes = ["Produit", "Unité", "Quantité vendue", "CA", "Marge"]
    for i, e in enumerate(entetes, start=1):
        c = ws2.cell(row=1, column=i, value=e)
        c.font, c.fill, c.border = blanc_gras, fond_bleu, bordure
    for r2, d in enumerate(detail_produits(debut, fin, boutique_id), start=2):
        for i, v in enumerate([d["nom"], d["unite"], d["quantite"],
                               d["ca"], d["marge"]], start=1):
            c = ws2.cell(row=r2, column=i, value=v)
            c.border = bordure
            if i >= 3:
                c.number_format = "#,##0.###" if i == 3 else "#,##0"
                c.alignment = droite

    # Feuille paiements
    ws3 = wb.create_sheet("Paiements")
    for i, e in enumerate(["Mode de paiement", "Nb ventes", "Montant"], start=1):
        c = ws3.cell(row=1, column=i, value=e)
        c.font, c.fill, c.border = blanc_gras, fond_bleu, bordure
    for r3, (mode, nb, total) in enumerate(
            repartition_paiements(debut, fin, boutique_id), start=2):
        for i, v in enumerate([mode, nb, total], start=1):
            c = ws3.cell(row=r3, column=i, value=v)
            c.border = bordure
            if i > 1:
                c.number_format = "#,##0"
                c.alignment = droite

    # Largeurs de colonnes
    for feuille in (ws, ws2, ws3):
        for col in range(1, 6):
            feuille.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================= PDF =================
def export_pdf(type_p: str, debut: date_type, fin: date_type,
               boutique_id: int | None = None,
               boutique_nom: str = "Toutes boutiques") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle("t", parent=styles["Title"],
                                 textColor=colors.HexColor("#" + BLEU))
    elements = [Paragraph("POISSONNERIE DONA — Cotonou", titre_style),
                Paragraph(titre_periode(type_p, debut, fin)
                          + f" — {boutique_nom}", styles["Heading2"]),
                Spacer(1, 6 * mm)]

    style_tbl = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BLEU)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ])

    ind = indicateurs_periode(debut, fin, boutique_id)
    elements.append(Table([
        ["Indicateur", "Valeur"],
        ["Chiffre d'affaires", _fmt_fcfa(ind["ca"])],
        ["Marge", _fmt_fcfa(ind["marge"])],
        ["Nombre de ventes", str(ind["nb_ventes"])],
        ["Volume vendu", f"{ind['volume']:g}"],
        ["Pertes (valeur)", _fmt_fcfa(ind["pertes"])],
    ], colWidths=[80 * mm, 60 * mm], style=style_tbl))
    elements.append(Spacer(1, 6 * mm))

    if type_p != "jour":
        data = [["Date", "CA", "Marge", "Ventes", "Pertes"]]
        for j in _jours_de(debut, fin):
            ij = indicateurs_periode(j, j, boutique_id)
            data.append([j.strftime("%d/%m/%Y"), _fmt_fcfa(ij["ca"]),
                         _fmt_fcfa(ij["marge"]), str(ij["nb_ventes"]),
                         _fmt_fcfa(ij["pertes"])])
        elements.append(Paragraph("Synthèse par jour", styles["Heading3"]))
        elements.append(Table(data, colWidths=[30 * mm, 35 * mm, 35 * mm,
                                               20 * mm, 35 * mm], style=style_tbl))
        elements.append(Spacer(1, 6 * mm))

    details = detail_produits(debut, fin, boutique_id)
    if details:
        data = [["Produit", "Qté", "CA", "Marge"]]
        for d in details:
            data.append([d["nom"], f"{d['quantite']:g} {d['unite']}",
                         _fmt_fcfa(d["ca"]), _fmt_fcfa(d["marge"])])
        elements.append(Paragraph("Détail par produit", styles["Heading3"]))
        elements.append(Table(data, colWidths=[55 * mm, 30 * mm, 35 * mm, 35 * mm],
                              style=style_tbl))

    doc.build(elements)
    return buf.getvalue()
